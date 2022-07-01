from datetime import timedelta, datetime
import openpyxl as opx
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from pathlib import Path

def create_report(list_trips, month):
    wb = opx.load_workbook(Path('excel/ts_temp.xlsx'))
    for id in list_trips:
        sheet = wb.active
        sh_ready = wb.copy_worksheet(sheet)
        thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        
        sh_ready.title = id[0]
        sh_ready['C11'].value = id[0]
        sh_ready['C12'].value = 'водитель' if id[4] else 'экспедитор'
        sh_ready['C2'].value = f"{id[3]}/{month}"
        sh_ready['D18'].value = id[2][0]
        sh_ready['D19'].value = id[2][1]
        sh_ready['G21'].value = id[0]
        for num, d_trips in enumerate(id[1]):
            row_ins = 17+num
            sh_ready.insert_rows(row_ins)
            for i in range(1, 8):
                sh_ready.cell(row=row_ins, column=i).border = thin_border
                sh_ready.cell(row=row_ins, column=i).alignment = Alignment(horizontal='center')
            sh_ready[f"A{row_ins}"].value = num + 1
            if d_trips[1] == 1:
                str_date = d_trips[0].split("-")
                sh_ready[f"B{row_ins}"].value = str_date[2] + '.' + str_date[1] + '.' + str_date[0]
            else:
                end_d=(datetime.strptime(d_trips[0], '%Y-%m-%d')+timedelta(days=int(d_trips[1])-1)).strftime('%d.%m.%Y')
                sh_ready[f"B{row_ins}"].value = d_trips[0].split('-')[2] + '-' + end_d
            sh_ready[f"C{row_ins}"].value = 'Доставка ТМЦ'
            sh_ready[f"D{row_ins}"].value = d_trips[1]
        
    wb.remove_sheet(sheet)
    wb.save(Path('excel/new.xlsx'))
    wb.close()

    